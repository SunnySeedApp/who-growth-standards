#!/usr/bin/env python3
"""Generate src/data/*.ts from the official WHO reference tables.

Downloads the WHO expanded tables (LMS by day for age-based indicators, by
length/height for the rest) from cdn.who.int and turns them into typed
TypeScript modules. Run this when the WHO publishes a revision:

    python3 scripts/generate-data.py

Downloads are cached in scripts/.cache, so re-runs are offline and cheap.
Requires: openpyxl (pip install openpyxl)
"""
import os
import urllib.request

BASE = ("https://cdn.who.int/media/docs/default-source/child-growth/"
        "child-growth-standards/indicators")

# indicator -> (url_path_template, x_axis, human_name)
SOURCES = {
    "wfa": ("weight-for-age/expanded-tables/wfa-{sex}-zscore-expanded-tables.xlsx",
            "day", "Weight-for-age"),
    "lhfa": ("length-height-for-age/expandable-tables/lhfa-{sex}-zscore-expanded-tables.xlsx",
             "day", "Length/height-for-age"),
    "hcfa": ("head-circumference-for-age/expanded-tables/hcfa-{sex}-zscore-expanded-tables.xlsx",
             "day", "Head-circumference-for-age"),
    "bfa": ("body-mass-index-for-age/expanded-tables/bfa-{sex}-zscore-expanded-tables.xlsx",
            "day", "BMI-for-age"),
    "wfl": ("weight-for-length-height/expanded-tables/wfl-{sex}-zscore-expanded-table.xlsx",
            "length", "Weight-for-length"),
    "wfh": ("weight-for-length-height/expanded-tables/wfh-{sex}-zscore-expanded-tables.xlsx",
            "height", "Weight-for-height"),
}

RAW = os.path.join(os.path.dirname(__file__), ".cache")
OUT = os.path.join(os.path.dirname(__file__), "..", "src", "data")


def fetch(indicator, sex):
    os.makedirs(RAW, exist_ok=True)
    path = os.path.join(RAW, f"{indicator}_{sex}.xlsx")
    if not os.path.exists(path):
        url = f"{BASE}/{SOURCES[indicator][0].format(sex=sex)}"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=90) as r, open(path, "wb") as f:
            f.write(r.read())
        print(f"  downloaded {indicator}_{sex}")
    return path


def read_lms(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    out = []
    for row in rows:
        if row[0] is None:
            continue
        out.append((float(row[0]), float(row[1]), float(row[2]), float(row[3])))
    return out


def emit(indicator, x_axis, human, data_by_sex):
    lines = [
        "// GENERATED FILE — do not edit by hand.",
        "// Source: WHO Child Growth Standards, expanded tables (LMS).",
        "// Regenerate with: python3 scripts/generate-data.py",
        f"// Indicator: {human}",
        "",
        'import type { LmsTable } from "../types";',
        "",
    ]
    for sex, rows in data_by_sex.items():
        start = rows[0][0]
        step = round(rows[1][0] - rows[0][0], 4)
        const = f"{indicator.upper()}_{sex.upper()}"
        lines.append(f"/** {human}, {sex}. x = {x_axis}, {start}…{rows[-1][0]}, step {step}. */")
        lines.append(f"export const {const}: LmsTable = {{")
        lines.append(f'  indicator: "{indicator}",')
        lines.append(f'  sex: "{"male" if sex == "boys" else "female"}",')
        lines.append(f'  xAxis: "{x_axis}",')
        lines.append(f"  start: {start},")
        lines.append(f"  step: {step},")
        lines.append("  lms: [")
        chunk = []
        for _, l, m, s in rows:
            chunk.append(f"[{l},{m},{s}]")
        for i in range(0, len(chunk), 8):
            lines.append("    " + ",".join(chunk[i:i + 8]) + ",")
        lines.append("  ],")
        lines.append("};")
        lines.append("")
    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, f"{indicator}.ts")
    with open(path, "w") as f:
        f.write("\n".join(lines))
    size = os.path.getsize(path) // 1024
    print(f"  wrote src/data/{indicator}.ts ({size} KB)")


def main():
    for indicator, (_, x_axis, human) in SOURCES.items():
        print(f"{indicator}:")
        data = {}
        for sex in ("boys", "girls"):
            data[sex] = read_lms(fetch(indicator, sex))
        emit(indicator, x_axis, human, data)


if __name__ == "__main__":
    main()
